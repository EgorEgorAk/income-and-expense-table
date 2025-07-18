from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook

def draw_tables(inc_category, inc_sum, exp_category, exp_sum, file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Бюджет"
    dwar1(ws, inc_category, inc_sum)
    expense(ws, exp_category, exp_sum)
    wb.save(file_path)

def dwar1(ws, inc_category, inc_sum):
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws['A2'] = '№'
    ws['B2'] = 'Категории доходов'
    ws['C2'] = 'Сумма'

    # Вставка пользовательских данных
    for i, (category, summ) in enumerate(zip(inc_category, inc_sum), start=3):

        ws[f'A{i}'] = i-2
        ws[f'B{i}'] = category
        ws[f'C{i}'] = summ
        ws[f'B{i}'].font = Font(name="Times New Roman", size=14)
        ws[f'C{i}'].font = Font(name="Times New Roman", size=14)

    last_row = 2 + len(inc_category)
    ws[f'B{last_row+1}'] = 'Итого'
    ws[f'C{last_row+1}'] = f'=SUM(C3:C{last_row})'

    # Оформление заголовка
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = "Доходы"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Таблица Excel
    table = Table(displayName="IncomeTable", ref=f"A2:C{last_row+1}")
    ws.add_table(table)
    thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    for row in ws[f"A2:C{last_row+1}"]:
        for cell in row:
            cell.border = thin_border

def expense(ws, exp_category, exp_sum):

    # Оформление и заголовки
    ws.merge_cells("A14:C14")
    ws["A14"].value = "Расходы"
    ws["A14"].font = Font(name="Times New Roman", size=18, bold=True, italic=True, underline="single")
    ws["A14"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A15"] = "№"
    ws["B15"] = "Категория расходов"
    ws["C15"] = "Сумма"
    ws["A15"].font = ws["B15"].font = ws["C15"].font = Font(name="Times New Roman", size=16, bold=True)
    ws["A15"].alignment = ws["B15"].alignment = ws["C15"].alignment = Alignment(horizontal="center", vertical="center")

    # Вставка пользовательских данных
    for i, (category, summ) in enumerate(zip(exp_category, exp_sum), start=16):
        ws.cell(row=i, column=1, value=i-15).font = Font(name="Times New Roman", size=14)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=category).font = Font(name="Times New Roman", size=14)
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="left")
        ws.cell(row=i, column=3, value=summ).font = Font(name="Times New Roman", size=14)
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="center")
    last_row = 15 + len(exp_category)

    ws[f'B{last_row+1}'] = 'Итого'
    ws[f'C{last_row+1}'] = f'=SUM(C16:C{last_row})'
    ws[f'B{last_row+1}'].font = ws[f'C{last_row+1}'].font = Font(name="Times New Roman", size=14, bold=True)


    # Таблица Excel

    table = Table(displayName="ExpenseTable", ref=f"A15:C{last_row+1}")
    ws.add_table(table)
    thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    for row in ws[f"A15:C{last_row+1}"]:
        for cell in row:
            cell.border = thin_border
def all(ws):
    ws.merge_cells("G1:I1")
    ws["G1"].value = "Отчет"
    ws["G1"].font = Font(name="Times New Roman", size=18, bold=True, italic=True, underline="single")
    ws["G"].alignment = Alignment(horizontal="center", vertical="center")
    ws["G2"] = "№"
    ws["H2"] = "Описание"
    ws["I2"] = "Сумма"
    ws["G3"] = 1
    ws["G4"] = 2
    ws["G5"] = 3
    ws["H2"] = "Доходы за месяц"
    ws["H3"] = "Расходы за месяц"
    ws["Н4"] = "Остаток"
    ws["I3"] = "=C8"
    ws["I4"] = "=C30"
    ws["I4"] = "=C8-C30"
    ws["A15"].font = ws["B15"].font = ws["C15"].font = Font(name="Times New Roman", size=16, bold=True)
    ws["A15"].alignment = ws["B15"].alignment = ws["C15"].alignment = Alignment(horizontal="center", vertical="center")


def fill_template(inc_category, inc_sum, exp_category, exp_sum, template_path, file_path):
    wb = load_workbook(template_path)
    ws = wb.active

    # Вставка доходов (B3:C7)
    for i, (category, summ) in enumerate(zip(inc_category, inc_sum), start=3):
        if i > 7:  # не больше строк, чем в шаблоне
            break
        ws[f'B{i}'] = category
        ws[f'C{i}'] = summ

    # Вставка расходов (B16:C29)
    for i, (category, summ) in enumerate(zip(exp_category, exp_sum), start=16):
        if i > 29:
            break
        ws[f'B{i}'] = category
        ws[f'C{i}'] = summ

    wb.save(file_path)