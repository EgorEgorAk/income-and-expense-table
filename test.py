from openpyxl import Workbook
from datetime import datetime

wb = Workbook()
ws = wb.active

# Дата
ws["A1"].value = datetime.now()
ws["A1"].number_format = "DD.MM.YYYY HH:MM"

# Деньги
ws["A2"].value = 1234.5
ws["A2"].number_format = "#,##0.00 ₽"

# Процент
ws["A3"].value = 0.82
ws["A3"].number_format = "0.00"

wb.save("example_formats.xlsx")